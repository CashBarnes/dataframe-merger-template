import pandas as pd
from openpyxl import load_workbook
import os

# Define input and output directories
input_dir = 'data'
output_dir = 'output'

# Define the column names to match on
data_source_col = ''  # Column name in the source DataFrame
data_target_col = ''  # Column name in the target DataFrame

data_source_id = ''  # ID column in the source DataFrame
data_target_id = ''  # ID column in the target DataFrame

# Initialize an empty DataFrame to store all merged data
all_data_df = pd.DataFrame()

# Process each file in the input directory
for file_name in os.listdir(input_dir):
    if file_name.endswith('.xlsx'):
        file_path = os.path.join(input_dir, file_name)

        # Load the data from the Excel file
        df = pd.read_excel(file_path)

        # Filter derived fields and non-derived fields
        derived_df = df[df[''] == '']  # Filter for derived fields
        non_derived_df = df[df[''] == '']  # Filter for non-derived fields

        # Merge the derived and non-derived DataFrames on the specified columns
        merged_df = pd.merge(
            derived_df,
            non_derived_df,
            how='left',
            left_on=[data_source_id, data_source_col],
            right_on=[data_target_id, data_target_col],
            suffixes=('_derived', '_non_derived')
        )

        # Define the final desired column order
        desired_columns = []  # List of columns in the desired order

        # Select and reorder the columns as per the desired structure
        final_df = merged_df[[col for col in desired_columns if col in merged_df.columns]]

        # Define output file paths
        output_file_name = f"{os.path.splitext(file_name)[0]}_merged_file_left_join"
        output_file_path_csv = os.path.join(output_dir, f"{output_file_name}.csv")
        output_file_path_xlsx = os.path.join(output_dir, f"{output_file_name}.xlsx")

        # Save the merged DataFrame to a CSV file
        final_df.to_csv(output_file_path_csv, index=False)

        # Save the merged DataFrame to an Excel file
        final_df.to_excel(output_file_path_xlsx, index=False)

        # Append the final DataFrame to the all_data_df
        all_data_df = pd.concat([all_data_df, final_df], ignore_index=True)

        # Adjust xlsx file so that it looks uniform when opening in spreadsheet editor
        wb = load_workbook(output_file_path_xlsx)
        ws = wb.active

        # Set the row height for all rows
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 15

        # Save the workbook
        wb.save(output_file_path_xlsx)

# Define the final output file path
final_output_file_path = os.path.join(output_dir, 'final_merged_output.csv')

# Save the concatenated DataFrame to a single CSV file
all_data_df.to_csv(final_output_file_path, index=False)

print("complete")